import json
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from flask import (
    Blueprint, render_template, request, jsonify, current_app, g, send_file
)
from .db import get_db

us = Blueprint('user', __name__)


# ─── Helper: send email notification ──────────────────────────────────────────

def send_email(to_email, subject, html_body):
    """Send email notification. Fails silently if not configured."""
    try:
        cfg = current_app.config
        if not cfg.get('MAIL_USERNAME') or not cfg.get('MAIL_PASSWORD'):
            current_app.logger.info(f"[EMAIL SKIPPED] To: {to_email} | Subject: {subject}")
            return False

        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = cfg['MAIL_USERNAME']
        msg['To'] = to_email
        msg.attach(MIMEText(html_body, 'html'))

        with smtplib.SMTP(cfg['MAIL_SERVER'], cfg['MAIL_PORT']) as server:
            server.starttls()
            server.login(cfg['MAIL_USERNAME'], cfg['MAIL_PASSWORD'])
            server.sendmail(cfg['MAIL_USERNAME'], to_email, msg.as_string())
        return True
    except Exception as e:
        current_app.logger.error(f"Email error: {e}")
        return False


def shift_email_html(action, employee_name, date, start_time, end_time, notes=''):
    """Generate HTML email body for shift notifications."""
    action_colors = {
        'created': '#059669',
        'updated': '#d97706',
        'deleted': '#dc2626'
    }
    action_labels = {
        'created': 'New Shift Assigned',
        'updated': 'Shift Updated',
        'deleted': 'Shift Cancelled'
    }
    color = action_colors.get(action, '#4f46e5')
    label = action_labels.get(action, 'Shift Notification')

    return f"""
    <div style="font-family: 'Segoe UI', sans-serif; max-width: 480px; margin: 0 auto; border-radius: 12px; overflow: hidden; border: 1px solid #e5e7eb;">
      <div style="background: {color}; padding: 24px; color: white;">
        <h2 style="margin: 0; font-size: 20px;">{label}</h2>
      </div>
      <div style="padding: 24px; background: #f9fafb;">
        <p style="margin: 0 0 16px; color: #374151;">Hi <strong>{employee_name}</strong>,</p>
        <p style="margin: 0 0 16px; color: #374151;">{"Your shift details are below:" if action != "deleted" else "The following shift has been cancelled:"}</p>
        <div style="background: white; border-radius: 8px; padding: 16px; border-left: 4px solid {color};">
          <p style="margin: 0 0 8px;"><strong>📅 Date:</strong> {date}</p>
          <p style="margin: 0 0 8px;"><strong>🕐 Start:</strong> {start_time}</p>
          <p style="margin: 0 0 8px;"><strong>🕔 End:</strong> {end_time}</p>
          {"<p style='margin: 0;'><strong>📝 Notes:</strong> " + notes + "</p>" if notes else ""}
        </div>
        <p style="margin: 16px 0 0; color: #6b7280; font-size: 13px;">This is an automated message from the Staff Rota System.</p>
      </div>
    </div>
    """


# ─── Pages ────────────────────────────────────────────────────────────────────

@us.route('/')
def index():
    return render_template('user/index.html')


@us.route('/rota')
def rota():
    return render_template('user/rota.html')


# ─── API: Employees ───────────────────────────────────────────────────────────

@us.route('/api/employees', methods=['GET'])
def get_employees():
    db = get_db()
    employees = db.execute(
        'SELECT id, name, email, role, color, hourly_pay FROM employees ORDER BY name'
    ).fetchall()
    return jsonify([dict(e) for e in employees])


@us.route('/api/employees', methods=['POST'])
def create_employee():
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'error': 'Name is required'}), 400

    db = get_db()
    cursor = db.execute(
        'INSERT INTO employees (name, email, role, color, hourly_pay) VALUES (?, ?, ?, ?, ?)',
        (data['name'], data.get('email', ''), data.get('role', 'Staff'),
         data.get('color', '#4f46e5'), float(data.get('hourly_pay', 0) or 0))
    )
    db.commit()
    emp = db.execute('SELECT * FROM employees WHERE id = ?', (cursor.lastrowid,)).fetchone()
    return jsonify(dict(emp)), 201


@us.route('/api/employees/<int:emp_id>', methods=['PUT'])
def update_employee(emp_id):
    data = request.get_json()
    db = get_db()
    db.execute(
        'UPDATE employees SET name=?, email=?, role=?, color=?, hourly_pay=? WHERE id=?',
        (data['name'], data.get('email', ''), data.get('role', 'Staff'),
         data.get('color', '#4f46e5'), float(data.get('hourly_pay', 0) or 0), emp_id)
    )
    db.commit()
    emp = db.execute('SELECT * FROM employees WHERE id = ?', (emp_id,)).fetchone()
    return jsonify(dict(emp))


@us.route('/api/employees/<int:emp_id>', methods=['DELETE'])
def delete_employee(emp_id):
    db = get_db()
    db.execute('DELETE FROM employees WHERE id = ?', (emp_id,))
    db.commit()
    return jsonify({'success': True})


# ─── API: Shifts ──────────────────────────────────────────────────────────────

@us.route('/api/shifts', methods=['GET'])
def get_shifts():
    db = get_db()
    start = request.args.get('start')
    end = request.args.get('end')

    query = '''
        SELECT s.id, s.employee_id, s.date, s.start_time, s.end_time,
               s.notes, s.calendar_event_id, s.created_at, s.updated_at,
               e.name AS employee_name, e.color AS employee_color, e.email AS employee_email, e.role AS employee_role
        FROM shifts s
        JOIN employees e ON s.employee_id = e.id
    '''
    params = []
    if start and end:
        query += ' WHERE s.date >= ? AND s.date <= ?'
        params = [start, end]
    query += ' ORDER BY s.date, s.start_time'

    shifts = db.execute(query, params).fetchall()
    return jsonify([dict(s) for s in shifts])


@us.route('/api/shifts', methods=['POST'])
def create_shift():
    data = request.get_json()
    required = ['employee_id', 'date', 'start_time', 'end_time']
    if not all(k in data for k in required):
        return jsonify({'error': 'Missing required fields'}), 400

    db = get_db()

    # Check employee exists
    emp = db.execute('SELECT * FROM employees WHERE id = ?', (data['employee_id'],)).fetchone()
    if not emp:
        return jsonify({'error': 'Employee not found'}), 404

    cursor = db.execute(
        '''INSERT INTO shifts (employee_id, date, start_time, end_time, notes)
           VALUES (?, ?, ?, ?, ?)''',
        (data['employee_id'], data['date'], data['start_time'], data['end_time'], data.get('notes', ''))
    )
    db.commit()
    shift_id = cursor.lastrowid

    shift = db.execute('''
        SELECT s.*, e.name AS employee_name, e.color AS employee_color,
               e.email AS employee_email, e.role AS employee_role
        FROM shifts s JOIN employees e ON s.employee_id = e.id
        WHERE s.id = ?
    ''', (shift_id,)).fetchone()
    shift_dict = dict(shift)

    # Send creation email
    if emp['email']:
        send_email(
            emp['email'],
            f"New Shift on {data['date']}",
            shift_email_html('created', emp['name'], data['date'], data['start_time'], data['end_time'], data.get('notes', ''))
        )

    return jsonify(shift_dict), 201


@us.route('/api/shifts/<int:shift_id>', methods=['PUT'])
def update_shift(shift_id):
    data = request.get_json()
    db = get_db()

    existing = db.execute('SELECT * FROM shifts WHERE id = ?', (shift_id,)).fetchone()
    if not existing:
        return jsonify({'error': 'Shift not found'}), 404

    emp_id = data.get('employee_id', existing['employee_id'])
    date = data.get('date', existing['date'])
    start_time = data.get('start_time', existing['start_time'])
    end_time = data.get('end_time', existing['end_time'])
    notes = data.get('notes', existing['notes'])

    db.execute(
        '''UPDATE shifts SET employee_id=?, date=?, start_time=?, end_time=?, notes=?,
           updated_at=CURRENT_TIMESTAMP WHERE id=?''',
        (emp_id, date, start_time, end_time, notes, shift_id)
    )
    db.commit()

    shift = db.execute('''
        SELECT s.*, e.name AS employee_name, e.color AS employee_color,
               e.email AS employee_email, e.role AS employee_role
        FROM shifts s JOIN employees e ON s.employee_id = e.id
        WHERE s.id = ?
    ''', (shift_id,)).fetchone()
    shift_dict = dict(shift)

    # Send update email
    emp = db.execute('SELECT * FROM employees WHERE id = ?', (emp_id,)).fetchone()
    if emp and emp['email']:
        send_email(
            emp['email'],
            f"Shift Updated – {date}",
            shift_email_html('updated', emp['name'], date, start_time, end_time, notes or '')
        )

    return jsonify(shift_dict)


def _calc_paid_hours(start_time, end_time):
    """
    Mirror of JS calcHours(). Returns (worked_mins, break_mins, paid_mins).
    Rules:
      - Strip trailing ≤30 min as unpaid break.
      - If total worked ≥ 8h, deduct an additional 1h.
    """
    try:
        sh, sm = map(int, start_time.split(':'))
        eh, em = map(int, end_time.split(':'))
        total = (eh * 60 + em) - (sh * 60 + sm)
        if total <= 0:
            return 0, 0, 0
        extra = total % 60
        brk = extra if 0 < extra <= 30 else 0
        if total >= 480:
            brk += 60
        return total, brk, total - brk
    except Exception:
        return 0, 0, 0


def _fmt_hours(mins):
    """Format minutes as 'XH:YYM' or 'XH'."""
    if mins == 0:
        return '0H'
    h, m = divmod(mins, 60)
    return f'{h}H:{m:02d}M' if m else f'{h}H'


@us.route('/api/shifts/export')
def export_shifts():
    """Download current view's shifts as a formatted .xlsx file."""
    db = get_db()
    start = request.args.get('start')
    end   = request.args.get('end')

    query = '''
        SELECT s.date, s.start_time, s.end_time, s.notes,
               e.name AS employee_name, e.role AS employee_role,
               COALESCE(e.hourly_pay, 0) AS hourly_pay
        FROM shifts s
        JOIN employees e ON s.employee_id = e.id
    '''
    params = []
    if start and end:
        query += ' WHERE s.date >= ? AND s.date <= ?'
        params = [start, end]
    query += ' ORDER BY s.date, s.start_time'
    shifts = db.execute(query, params).fetchall()

    # ── Pre-calculate per-shift pay, group by employee ────────────────────
    from collections import defaultdict, OrderedDict

    # employee_name → {'rate': float, 'shifts': int, 'paid_mins': int, 'total_pay': float}
    summary = OrderedDict()
    shift_rows = []  # enriched rows for the detail sheet

    for s in shifts:
        _, brk, paid_mins = _calc_paid_hours(s['start_time'], s['end_time'])
        paid_hrs  = paid_mins / 60
        rate      = float(s['hourly_pay'] or 0)
        shift_pay = round(paid_hrs * rate, 2)

        name = s['employee_name']
        if name not in summary:
            summary[name] = {'rate': rate, 'shifts': 0, 'paid_mins': 0, 'total_pay': 0.0}
        summary[name]['shifts']    += 1
        summary[name]['paid_mins'] += paid_mins
        summary[name]['total_pay'] += shift_pay

        shift_rows.append({
            'name': name, 'date': s['date'],
            'start': s['start_time'], 'end': s['end_time'],
            'break': _fmt_hours(brk) if brk else '–',
            'paid_hrs': round(paid_hrs, 2),
            'rate': rate, 'pay': shift_pay,
        })

    # ── Shared styles ─────────────────────────────────────────────────────
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='6366F1')
    totals_font = Font(bold=True, size=11)
    totals_fill = PatternFill('solid', fgColor='E0E7FF')
    alt_fill    = PatternFill('solid', fgColor='F4F4FF')
    center      = Alignment(horizontal='center', vertical='center')
    border      = Border(
        bottom=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin',  color='D1D5DB'),
    )
    gbp_fmt = u'\u00a3#,##0.00'

    def style_header_row(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = border
        ws.row_dimensions[1].height = 22

    def apply_cell(ws, row, col, value=None, fmt=None, fill=None, font=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = center
        cell.border    = border
        if fmt:  cell.number_format = fmt
        if fill: cell.fill = fill
        if font: cell.font = font
        return cell

    wb = openpyxl.Workbook()

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 1 — Payroll Summary (one row per employee)
    # ═══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Payroll Summary'

    sum_headers = ['Employee', 'Shifts', 'Total Hours', 'Hourly Rate', 'Total Pay']
    ws1.append(sum_headers)
    style_header_row(ws1, len(sum_headers))

    grand_mins = 0
    grand_pay  = 0.0

    for row_idx, (name, d) in enumerate(summary.items(), 2):
        hrs = d['paid_mins'] / 60
        fill = alt_fill if row_idx % 2 == 0 else None
        apply_cell(ws1, row_idx, 1, name,              fill=fill)
        apply_cell(ws1, row_idx, 2, d['shifts'],       fill=fill)
        apply_cell(ws1, row_idx, 3, round(hrs, 2),     fill=fill)
        apply_cell(ws1, row_idx, 4, d['rate'],         fill=fill, fmt=gbp_fmt)
        apply_cell(ws1, row_idx, 5, round(d['total_pay'], 2), fill=fill, fmt=gbp_fmt)
        grand_mins += d['paid_mins']
        grand_pay  += d['total_pay']

    # Totals row
    tr = len(summary) + 2
    apply_cell(ws1, tr, 1, 'TOTAL',                      fill=totals_fill, font=totals_font)
    apply_cell(ws1, tr, 2, sum(d['shifts'] for d in summary.values()),
               fill=totals_fill, font=totals_font)
    apply_cell(ws1, tr, 3, round(grand_mins / 60, 2),   fill=totals_fill, font=totals_font)
    apply_cell(ws1, tr, 4, '',                           fill=totals_fill, font=totals_font)
    apply_cell(ws1, tr, 5, round(grand_pay, 2),          fill=totals_fill, font=totals_font, fmt=gbp_fmt)

    for w, col in zip([24, 10, 14, 14, 14], range(1, 6)):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 2 — Shift Detail (one row per shift, for reference)
    # ═══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Shift Detail')

    det_headers = ['Employee', 'Date', 'Start', 'End', 'Break', 'Paid Hours', 'Hourly Rate', 'Shift Pay']
    ws2.append(det_headers)
    style_header_row(ws2, len(det_headers))

    for row_idx, r in enumerate(shift_rows, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        apply_cell(ws2, row_idx, 1, r['name'],     fill=fill)
        apply_cell(ws2, row_idx, 2, r['date'],     fill=fill)
        apply_cell(ws2, row_idx, 3, r['start'],    fill=fill)
        apply_cell(ws2, row_idx, 4, r['end'],      fill=fill)
        apply_cell(ws2, row_idx, 5, r['break'],    fill=fill)
        apply_cell(ws2, row_idx, 6, r['paid_hrs'], fill=fill)
        apply_cell(ws2, row_idx, 7, r['rate'],     fill=fill, fmt=gbp_fmt)
        apply_cell(ws2, row_idx, 8, r['pay'],      fill=fill, fmt=gbp_fmt)

    for w, col in zip([22, 13, 10, 10, 10, 13, 14, 13], range(1, 9)):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # Stream to response
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"rota_{start or 'all'}_to_{end or 'all'}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@us.route('/api/shifts/<int:shift_id>', methods=['DELETE'])
def delete_shift(shift_id):
    db = get_db()

    shift = db.execute('''
        SELECT s.*, e.name AS employee_name, e.email AS employee_email
        FROM shifts s JOIN employees e ON s.employee_id = e.id
        WHERE s.id = ?
    ''', (shift_id,)).fetchone()

    if not shift:
        return jsonify({'error': 'Shift not found'}), 404

    shift_dict = dict(shift)
    db.execute('DELETE FROM shifts WHERE id = ?', (shift_id,))
    db.commit()

    # Send cancellation email
    if shift_dict.get('employee_email'):
        send_email(
            shift_dict['employee_email'],
            f"Shift Cancelled – {shift_dict['date']}",
            shift_email_html('deleted', shift_dict['employee_name'],
                             shift_dict['date'], shift_dict['start_time'], shift_dict['end_time'])
        )

    return jsonify({'success': True, 'id': shift_id})
